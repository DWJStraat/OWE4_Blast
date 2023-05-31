"""
This file is used to parse fastq files and return the data in a format that can be used by the database.
Created on 10-may-2021 by David
Collaborators: David
Last modified on 31-may-2023 by David
"""
import openpyxl
import json


class Excel:
    """
    This class is used to parse fastq files and return the data in a format that can be used by the database.
    ::param excel_sheet: The excel sheet to be parsed. Has to be a single page.
    """

    def __init__(self, excel_sheet):
        self.values = None
        self.db_data = None
        if type(excel_sheet) is str:
            self.excel = open(excel_sheet, "r")
            self.excel_name = excel_sheet
        else:
            self.excel = excel_sheet
            self.excel_name = excel_sheet.name
        self.df = openpyxl.load_workbook(self.excel_name)
        self.sheet = self.df.active
        self.entries = []
        self.parse()

    def parse(self):
        """
        This method parses the excel sheet and returns the data in a format that can be used by the database.
        """
        for row in self.sheet.iter_rows():
            val = json.loads('{}')
            val["fwd_header"] = row[0].value
            val["fwd_seq"] = row[1].value
            val["fwd_qual"] = row[2].value
            val["rev_header"] = row[3].value
            val["rev_seq"] = row[4].value
            val["rev_qual"] = row[5].value
            self.entries.append(val)

    def parse_for_db(self, use_json=True):
        """
        This method parses the excel sheet and returns the data in a format that can be used by the database.
        :param use_json: whether to use json or not
        """
        data = json.loads('{}')
        values = []
        value_id = 0
        for row in self.sheet.iter_rows():
            if use_json:
                data[value_id] = json.loads('{}')
                data[value_id]["header"] = row[0].value
                data[value_id]["sequence"] = row[1].value
                data[value_id]["quality"] = row[2].value
                value_id += 1
                data[value_id] = json.loads('{}')
                data[value_id]["header"] = row[3].value
                data[value_id]["sequence"] = row[4].value
                data[value_id]["quality"] = row[5].value
                value_id += 1
            else:
                values += (row[0].value, row[1].value, row[2].value), (row[3].value, row[4].value, row[5].value)
        self.db_data = data
        self.values = values


if __name__ == "__main__":
    a = Excel("Map1.xlsx")
    a.parse_for_db(use_json=False)
